import os
import json
from conftest import testdata_dir
import openpyxl
# import pytest

excel_dir = os.path.join(testdata_dir, "_excel")
excel_files = [
    os.path.join(excel_dir, "test-default.xlsx"),
    os.path.join(excel_dir, "test-data-row.xlsx"),
]


# @pytest.fixture(scope="module")
def setup_module():
    os.makedirs(excel_dir, exist_ok=True)
    pass


def F(fname):
    return os.path.join(excel_dir, fname)


def test_default():
    gen_default()
    from excel2xx import main

    argv = ["json", "-v", "-o", F("yes.json"), excel_files[0]]
    code = main.main_docopt(argv)
    assert 0 == code
    with open(F("yes.json")) as fp:
        data = json.load(fp)
        pass
    assert 3 == len(data["auto"])
    assert 1 == len(data["sheet0"])
    assert 1 == len(data["sheet1"])
    assert 1 == len(data["sheet2"])
    assert 3 == len(data["sheet-map"])
    pass


def test_datarow():
    gen_datarow()
    from excel2xx import main

    argv = ["json", "--data-row", "6", "-o", F("yes.json"), excel_files[1]]
    code = main.main_docopt(argv)
    assert 0 == code
    with open(F("yes.json")) as fp:
        data = json.load(fp)
        pass
    assert 3 == len(data["auto"])
    assert 1 == len(data["sheet0"])
    assert 1 == len(data["sheet1"])
    assert 1 == len(data["sheet2"])
    assert 3 == len(data["sheet-map"])
    pass


def gen_default():
    setup_module()
    w = openpyxl.Workbook()
    s = w.create_sheet(title="auto")
    s.append(["key", "type", "value"])
    s.append(["string", "auto", "auto"])
    s.append(["", "", ""])
    s.append(["key1", "int", 1])
    s.append(["key2", "string", "2"])
    s.append(["key3", "array<int>", "1,2,3"])
    for i in range(3):
        s = w.create_sheet(title=f"sheet{i}")
        s.append(["key1", "key2", "key3", "key3"])
        s.append(["int", "string", "float", "array<int>"])
        s.append(["", "", "", ""])
        s.append([1, "2", 3.0, "1,2,3"])
        pass

    s = w.create_sheet("sheet-map(map)")
    s.append(["key1", "key2"])
    s.append(["int", "string"])
    s.append(["", ""])
    s.append([1, "1"])
    s.append([2, "2"])
    s.append([3, "3"])

    w.save(excel_files[0])
    pass


def gen_datarow():
    setup_module()
    w = openpyxl.Workbook()
    s = w.create_sheet(title="auto")
    s.append(["key", "type", "value"])
    s.append(["string", "auto", "auto"])
    s.append(["", "", ""])
    s.append(["", "", ""])
    s.append(["", "", ""])
    s.append(["key1", "int", 1])
    s.append(["key2", "string", "2"])
    s.append(["key3", "array<int>", "1,2,3"])
    for i in range(3):
        s = w.create_sheet(title=f"sheet{i}")
        s.append(["key1", "key2", "key3", "key3"])
        s.append(["int", "string", "float", "array<int>"])
        s.append(["", "", "", ""])
        s.append(["", "", "", ""])
        s.append(["", "", "", ""])

        s.append([1, "2", 3.0, "1,2,3"])
        pass

    s = w.create_sheet("sheet-map(map)")
    s.append(["key1", "key2"])
    s.append(["int", "string"])
    s.append(["", ""])
    s.append(["", ""])
    s.append(["", ""])
    s.append([1, "1"])
    s.append([2, "2"])
    s.append([3, "3"])

    w.save(excel_files[1])
    pass
